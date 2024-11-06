# import csv

# image_data = [
#     ["product1", "posters(1080x1920)/resized_...First_Do_No_Harm.jpg"],
#     ["product2", "posters(1080x1920)/resized_xXx_Return_of_Xander_Cage.jpg"],
#     ["product3", "posters(1080x1920)/resized_xXx.jpg"]
# ]

# with open("products.csv", "w", newline="") as file:
#     writer = csv.writer(file)
#     writer.writerows(image_data)


import openpyxl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import io
import base64

def embed_image_in_xlsx(xlsx_file, sheet_name, cell, image_path):
    # Open or create the workbook
    try:
        workbook = openpyxl.load_workbook(xlsx_file)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    
    # Select the sheet or create a new one
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)
    
    # Open and resize the image
    with Image.open(image_path) as img:
        img.thumbnail((200, 200))  # Resize image to fit in cell
        
        # Save image to a bytes buffer
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        
        # Create openpyxl image
        xl_image = XLImage(buffer)
        
        workbook.active.add_image(xl_image, cell)
        # Add image to the specified cell
        # sheet.add_image(xl_image, cell)
    
    # Save the workbook
    workbook.save(xlsx_file)

# Example usage:
embed_image_in_xlsx('products.xlsx', "sheet1", 'A1', 'posters(1080x1920)/resized_...First_Do_No_Harm.jpg')
